import requests
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import io
import os
import time

# Importar openpyxl diretamente em vez de usar pandas
try:
    import openpyxl
    USE_OPENPYXL = True
except ImportError:
    # Fallback para pandas se openpyxl não estiver disponível
    USE_OPENPYXL = False
    try:
        import pandas as pd
    except ImportError:
        logging.error("Nem openpyxl nem pandas estão disponíveis. Não será possível processar arquivos Excel.")

# Configuração de logging
logger = logging.getLogger('sharepoint_client')

class SharePointClient:
    """Cliente para acessar dados de planilhas do SharePoint/OneDrive"""

    def __init__(self, public_url: str):
        self.public_url = public_url
        self._cache = {}
        self._cache_expiry = {}
        self.cache_duration = 300  # 5 minutos

    def _download_file(self, url: str) -> Optional[bytes]:
        """Baixa o arquivo da URL fornecida com tratamento robusto de erros"""
        max_retries = 3
        retry_delay = 2  # segundos
        
        for attempt in range(1, max_retries + 1):
            try:
                # Usar headers mais completos para simular um navegador real
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                    'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/vnd.ms-excel, */*',
                    'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7',
                    'Cache-Control': 'no-cache',
                    'Pragma': 'no-cache'
                }
                
                logger.info(f"Tentativa {attempt}/{max_retries}: Iniciando download do arquivo: {url}")
                
                # Adicionar parâmetros para evitar cache
                from datetime import datetime
                import time
                timestamp = int(time.time() * 1000)
                url_with_nocache = f"{url}&_={timestamp}" if '?' in url else f"{url}?_={timestamp}"
                
                response = requests.get(url_with_nocache, headers=headers, timeout=60, allow_redirects=True)
                response.raise_for_status()
                
                # Verificar informações da resposta
                content_type = response.headers.get('Content-Type', '')
                content_length = len(response.content)
                logger.info(f"Download concluído. Content-Type: {content_type}, Tamanho: {content_length} bytes")
                
                # Verificar se o conteúdo parece ser HTML (possível página de erro ou redirecionamento)
                if content_type.startswith('text/html') or (len(response.content) > 10 and response.content[:10].decode('utf-8', errors='ignore').strip().startswith('<')):
                    logger.warning(f"O servidor retornou HTML em vez de um arquivo Excel. Tentativa {attempt}/{max_retries}")
                    if attempt < max_retries:
                        logger.info(f"Aguardando {retry_delay} segundos antes da próxima tentativa...")
                        time.sleep(retry_delay)
                        retry_delay *= 2  # Aumentar o tempo de espera entre tentativas
                        continue
                    else:
                        logger.error("Todas as tentativas falharam. O servidor continua retornando HTML.")
                        # Salvar o HTML para diagnóstico
                        try:
                            with open('download_error.html', 'wb') as f:
                                f.write(response.content)
                            logger.info("Conteúdo HTML salvo em 'download_error.html' para diagnóstico")
                        except Exception as save_err:
                            logger.error(f"Não foi possível salvar o HTML para diagnóstico: {save_err}")
                        return None
                
                # Verificar se o arquivo parece ser um arquivo Excel válido (começa com PK - assinatura ZIP)
                if not response.content.startswith(b'PK'):
                    logger.warning(f"O conteúdo baixado não parece ser um arquivo Excel válido. Tentativa {attempt}/{max_retries}")
                    if attempt < max_retries:
                        logger.info(f"Aguardando {retry_delay} segundos antes da próxima tentativa...")
                        time.sleep(retry_delay)
                        retry_delay *= 2
                        continue
                    else:
                        logger.error("Todas as tentativas falharam. O conteúdo não é um arquivo Excel válido.")
                        # Salvar o conteúdo para diagnóstico
                        try:
                            with open('invalid_excel.bin', 'wb') as f:
                                f.write(response.content)
                            logger.info("Conteúdo inválido salvo em 'invalid_excel.bin' para diagnóstico")
                        except Exception as save_err:
                            logger.error(f"Não foi possível salvar o conteúdo para diagnóstico: {save_err}")
                        return None
                
                # Se chegou aqui, o download foi bem-sucedido
                return response.content
                
            except requests.RequestException as e:
                logger.error(f"Erro ao baixar arquivo (tentativa {attempt}/{max_retries}): {e}")
                if attempt < max_retries:
                    logger.info(f"Aguardando {retry_delay} segundos antes da próxima tentativa...")
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    logger.error("Todas as tentativas de download falharam.")
                    return None
        
        return None  # Não deveria chegar aqui, mas por segurança

    def _get_cached_data(self, cache_key: str) -> Optional[any]:
        if cache_key in self._cache:
            if datetime.now() < self._cache_expiry[cache_key]:
                return self._cache[cache_key]
            del self._cache[cache_key]
            del self._cache_expiry[cache_key]
        return None

    def _set_cache_data(self, cache_key: str, data: any):
        self._cache[cache_key] = data
        self._cache_expiry[cache_key] = datetime.now() + timedelta(seconds=self.cache_duration)

    def get_placas_data(self, force_refresh: bool = False) -> List[Dict[str, str]]:
        """Obtém dados de placas da planilha do SharePoint com tratamento robusto de erros"""
        global USE_OPENPYXL
        cache_key = 'placas_data'

        if not force_refresh:
            cached_data = self._get_cached_data(cache_key)
            if cached_data is not None:
                logger.debug("Dados das placas obtidos do cache")
                return cached_data

        try:
            logger.info("Baixando planilha de placas do SharePoint/OneDrive...")
            file_content = self._download_file(self.public_url)
            if file_content is None:
                logger.error("Não foi possível baixar a planilha após várias tentativas")
                # Tentar usar dados em cache mesmo que estejam expirados
                if cache_key in self._cache:
                    logger.warning("Usando dados em cache expirados como fallback")
                    return self._cache[cache_key]
                return []

            # O método _download_file já verifica se o arquivo é um ZIP válido
            # então podemos prosseguir diretamente para o processamento
            
            excel_file = io.BytesIO(file_content)
            try:
                # Usar um bloco try específico para a leitura do Excel
                logger.info("Processando arquivo Excel...")
                
                placas_data = []
                
                # Usar openpyxl diretamente em vez de pandas
                if USE_OPENPYXL:
                    try:
                        logger.info("Usando openpyxl para processar o arquivo Excel")
                        workbook = openpyxl.load_workbook(excel_file, read_only=True)
                        
                        # Verificar se a aba 'PLACAS' existe
                        if 'PLACAS' in workbook.sheetnames:
                            sheet = workbook['PLACAS']
                            
                            # Encontrar o índice da coluna 'PLACA'
                            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                            placa_col_idx = None
                            
                            for i, cell_value in enumerate(header_row):
                                if cell_value and str(cell_value).upper().strip() == 'PLACA':
                                    placa_col_idx = i
                                    break
                            
                            if placa_col_idx is not None:
                                # Coletar todas as placas da coluna
                                placas_set = set()
                                for row in sheet.iter_rows(min_row=2, values_only=True):
                                    if len(row) > placa_col_idx and row[placa_col_idx]:
                                        placa = str(row[placa_col_idx]).strip()
                                        if placa:
                                            placas_set.add(placa.upper())
                                
                                # Converter para o formato esperado
                                for placa in sorted(placas_set):
                                    placas_data.append({
                                        'placa': placa,
                                        'label': placa,
                                        'value': placa
                                    })
                                logger.info(f"Carregadas {len(placas_data)} placas da planilha usando openpyxl")
                            else:
                                logger.warning("Coluna 'PLACA' não encontrada na aba 'PLACAS'")
                        else:
                            logger.warning("Aba 'PLACAS' não encontrada na planilha")
                        
                        # Fechar o workbook para liberar recursos
                        workbook.close()
                    except Exception as openpyxl_err:
                        logger.error(f"Erro ao processar Excel com openpyxl: {openpyxl_err}")
                        # Tentar com pandas como fallback
                        USE_OPENPYXL = False
                        # Resetar o ponteiro do arquivo para o início
                        excel_file.seek(0)
                
                # Usar pandas como fallback se openpyxl falhar ou não estiver disponível
                if not USE_OPENPYXL and len(placas_data) == 0:
                    try:
                        import pandas as pd
                        logger.info("Usando pandas como fallback para processar o arquivo Excel")
                        # Resetar o ponteiro do arquivo para o início
                        excel_file.seek(0)
                        df = pd.read_excel(excel_file, sheet_name='PLACAS')
                        
                        if 'PLACA' in df.columns:
                            # Processar coluna de placas
                            placas_validas = df['PLACA'].dropna().astype(str).str.strip()
                            placas_validas = placas_validas[placas_validas != ''].unique()
                            
                            # Normalizar e ordenar placas
                            for placa in sorted(placas_validas):
                                placa_normalizada = placa.upper().strip()
                                placas_data.append({
                                    'placa': placa_normalizada,
                                    'label': placa_normalizada,
                                    'value': placa_normalizada
                                })
                            logger.info(f"Carregadas {len(placas_data)} placas da planilha usando pandas")
                        else:
                            logger.warning("Coluna 'PLACA' não encontrada na aba 'PLACAS' (pandas)")
                    except Exception as pandas_err:
                        logger.error(f"Erro ao processar Excel com pandas: {pandas_err}")
                        # Ambos os métodos falharam
                
                # Atualizar cache apenas se processamento foi bem-sucedido e encontramos placas
                if len(placas_data) > 0:
                    logger.info(f"Total de {len(placas_data)} placas carregadas com sucesso")
                    self._set_cache_data(cache_key, placas_data)
                    return placas_data
                else:
                    logger.warning("Coluna 'PLACA' não encontrada na aba 'PLACAS'")
                    # Tentar usar dados em cache mesmo que estejam expirados
                    if cache_key in self._cache:
                        logger.warning("Usando dados em cache expirados como fallback")
                        return self._cache[cache_key]
                    return []
                    
            except Exception as excel_err:
                logger.error(f"Erro ao processar o arquivo Excel: {excel_err}")
                # Salvar o arquivo para diagnóstico
                try:
                    with open('excel_error.xlsx', 'wb') as f:
                        f.write(file_content)
                    logger.info("Arquivo Excel salvo como 'excel_error.xlsx' para diagnóstico")
                except Exception as save_err:
                    logger.error(f"Não foi possível salvar o arquivo Excel para diagnóstico: {save_err}")
                
                # Tentar usar dados em cache mesmo que estejam expirados
                if cache_key in self._cache:
                    logger.warning("Usando dados em cache expirados como fallback após erro de processamento")
                    return self._cache[cache_key]
                return []
                
        except Exception as e:
            logger.error(f"Erro geral ao processar planilha de placas: {e}")
            # Tentar usar dados em cache mesmo que estejam expirados
            if cache_key in self._cache:
                logger.warning("Usando dados em cache expirados como fallback após erro geral")
                return self._cache[cache_key]
            return []

    def get_placas_list(self, force_refresh: bool = False) -> List[str]:
        """Retorna a lista de placas da planilha do SharePoint"""
        try:
            placas_data = self.get_placas_data(force_refresh)
            if not placas_data:
                logger.warning("Nenhuma placa encontrada na planilha")
                return []
            
            placas = [placa['placa'] for placa in placas_data if 'placa' in placa and placa['placa']]
            logger.info(f"Retornando lista de {len(placas)} placas")
            return placas
        except Exception as e:
            logger.error(f"Erro ao obter lista de placas: {e}")
            return []

    def refresh_cache(self) -> bool:
        try:
            placas_data = self.get_placas_data(force_refresh=True)
            return len(placas_data) > 0
        except Exception as e:
            logger.error(f"Erro ao atualizar cache: {e}")
            return False

    def get_cache_status(self) -> Dict[str, any]:
        cache_key = 'placas_data'
        if cache_key in self._cache:
            return {
                'cached': True,
                'expires_at': self._cache_expiry[cache_key].isoformat(),
                'items_count': len(self._cache[cache_key]),
                'seconds_to_expiry': (self._cache_expiry[cache_key] - datetime.now()).total_seconds()
            }
        else:
            return {
                'cached': False,
                'expires_at': None,
                'items_count': 0,
                'seconds_to_expiry': 0
            }

# Instância global do cliente
sharepoint_client = None

def get_sharepoint_client() -> SharePointClient:
    global sharepoint_client
    if sharepoint_client is None:
        # URL atualizada do OneDrive conforme fornecido pelo usuário
        # Importante: usar o parâmetro download=1 para obter o arquivo diretamente
        base_url = "https://ictsi-my.sharepoint.com/:x:/p/leonardo_fragoso_itracker/EV8B0yiu9txKjo3I45WIYXkBK9u7ye7q9YF7bMb1E81fOA"
        
        # Garantir que estamos usando o parâmetro download=1 em vez de e=7tWqt3
        if "?" in base_url:
            # Remover qualquer parâmetro existente e adicionar download=1
            public_url = base_url.split("?")[0] + "?download=1"
        else:
            public_url = base_url + "?download=1"
            
        logger.info(f"Inicializando SharePointClient com URL base: {public_url}")
        sharepoint_client = SharePointClient(public_url)
    return sharepoint_client

def get_placas_list(force_refresh: bool = False) -> List[str]:
    client = get_sharepoint_client()
    return client.get_placas_list(force_refresh)

def refresh_placas_cache() -> bool:
    client = get_sharepoint_client()
    return client.refresh_cache()

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    print("Testando cliente SharePoint...")
    client = get_sharepoint_client()
    placas = client.get_placas_list()
    print(f"Encontradas {len(placas)} placas:")
    for i, placa in enumerate(placas[:10]):
        print(f"  {i+1}. {placa}")
    if len(placas) > 10:
        print(f"  ... e mais {len(placas) - 10} placas")
    print("\nStatus do cache:")
    cache_status = client.get_cache_status()
    print(f"  Cache ativo: {cache_status['cached']}")
    print(f"  Itens em cache: {cache_status['items_count']}")
